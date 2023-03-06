import sqlite3
from models import Student
from openpyxl import Workbook
from datetime import date


class DataBase ():
    def __init__(self, filename: str = 'uwc_database.db'):
        self.connection = sqlite3.connect(filename, check_same_thread=False)

    def create_table_users(self):
        self.connection.execute("""
            CREATE TABLE users (
                id INTEGER PRIMARY KEY,
                telegram_id INT NOT NULL,
                name VARCHAR(50) NOT NULL,
                surname VARCHAR(50),
                college VARCHAR(50),
                year_start INTEGER,
                year_finish INTEGER,
                email VARCHAR(50),
                social_network VARCAHR,
                comms_chanel VARCHAR,
                share_info BOOL,
                live_place VARCHAR(70),
                university VARCHAR(100),
                work VARCHAR,
                interests VARCAHR(100),
                update_date DATE,
                is_admin BOOL
            );
        """)

        self.connection.commit()

    def create_table_colleges(self):
        self.connection.execute("""
            CREATE TABLE colleges (
                id INTEGER PRIMARY KEY,
                name VARCHAR(100) NOT NULL,
                location VARHAR(50) NOT NULL
            );
        """)

        self.connection.commit()

    def add_admin_to_db(self, telegram_id: int, name: str):
        self.connection.execute("""
            INSERT INTO users (telegram_id, name, is_admin) VALUES (?, ?, True)
        """, (telegram_id, name))
        self.connection.commit()

    def remove_admin_from_db(self, telegram_id: int):
        cursor = self.connection.cursor()

        cursor.execute("""
            DELETE FROM users
            WHERE telegram_id = ?
            AND is_admin = True
            """, (telegram_id,))

        self.connection.commit()

    def show_admin_from_db(self):
        pass

    def is_admin(self, telegram_id: int) -> bool:
        cursor = self.connection.cursor()

        cursor.execute(f"""
            SELECT EXISTS(SELECT * FROM users
            WHERE is_admin = True
            AND telegram_id = {telegram_id})
            """)

        return cursor.fetchone()[0]

    def add_user_to_db(self, student: Student):
        cursor = self.connection.cursor()

        cursor.execute("""
            INSERT INTO users (
                telegram_id,
                name,
                surname,
                college,
                year_start,
                year_finish,
                email,
                social_network,
                comms_chanel,
                share_info,
                live_place,
                university,
                work,
                interests
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, student.get_info())

        cursor.execute("""
            UPDATE users
            SET update_date = ?, is_admin = ?
            WHERE telegram_id = ?
        """, (date.today(), False, student.telegram_id))

        self.connection.commit()

    def update_user_in_db(self, student: Student):
        cursor = self.connection.cursor()

        cursor.execute("""
            UPDATE users
            SET
                email = ?,
                live_place = ?,
                work = ?
            WHERE telegram_id = ?
        """, (
            student.email, student.live_place,
            student.work, student.telegram_id
        )
        )

        db.refresh_update_date(student.telegram_id)

        self.connection.commit()

    def refresh_update_date(self, telegram_id: int):
        cursor = self.connection.cursor()

        cursor.execute("""
            UPDATE users
            SET update_date = ?
            WHERE telegram_id = ?
        """, (date.today(), telegram_id))

        self.connection.commit()

    def get_telegram_id_list(self) -> list:
        cursor = self.connection.cursor()

        cursor.execute("""
            SELECT telegram_id
            FROM users
            WHERE is_admin = False
        """)
        temp = cursor.fetchall()

        user_list = set([x[0] for x in temp])

        return user_list

    def show_all_users(self):
        pass

    def get_student_info(self, telegram_id: int) -> Student:
        cursor = self.connection.cursor()

        cursor.execute(f"""
            SELECT
                telegram_id,
                name,
                surname,
                college,
                year_start,
                year_finish,
                email,
                social_network,
                comms_chanel,
                share_info,
                live_place,
                university,
                work,
                interests,
                update_date
            FROM users
            WHERE telegram_id = {telegram_id}
        """)

        student = Student(*cursor.fetchone())

        return student

    def if_update_need(self, telegram_id):
        pass

    def export_student_to_excell(self):
        cursor = self.connection.cursor()
        cursor.execute("SELECT * FROM users")

        rows = cursor.fetchall()

        wb = Workbook()
        ws = wb.active

        headers = [description[0] for description in cursor.description]
        ws.append(headers)

        for row in rows:
            ws.append(row)

        wb.save('uwc_members.xlsx')

    def export_admins_to_excell(self):
        cursor = self.connection.cursor()
        cursor.execute("""
            SELECT telegram_id, name FROM users
            WHERE is_admin = True
            """)

        rows = cursor.fetchall()

        wb = Workbook()
        ws = wb.active

        headers = [description[0] for description in cursor.description]
        ws.append(headers)

        for row in rows:
            ws.append(row)

        wb.save('uwc_admins.xlsx')

    def export_colleges_to_excell(self):
        cursor = self.connection.cursor()
        cursor.execute("SELECT * FROM colleges")

        rows = cursor.fetchall()

        wb = Workbook()
        ws = wb.active

        headers = [description[0] for description in cursor.description]
        ws.append(headers)

        for row in rows:
            ws.append(row)

        wb.save('uwc_colleges.xlsx')


db = DataBase()
